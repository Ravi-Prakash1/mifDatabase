#!/usr/bin/env python3
"""
fetch_news.py — pull fresh market news into MIF_News_Intelligence.xlsx.

Pipeline (all sources free):
  1. Fetch candidate headlines from Google News RSS (topic + company queries)
     and a couple of industry RSS feeds.
  2. De-duplicate against data/news_seen.json so nothing is published twice.
  3. Fetch the USD/INR spot rate from a free, no-key FX endpoint.
  4. Summarise via a FREE engine (Groq/Gemini) — or, with no key, rule-based —
     relevant to MIF, verify them, and shape them into the exact sheet schema
     that generate_portal_v7_News.py::parse_news() reads. Every story keeps a
     real Source URL — unsourced items are dropped.
  5. Append one dated edition (Briefing / Indicators / Trend / News rows) to the
     workbook and record the new story keys in the seen cache.

Nothing here deploys anything: the workflow opens a PR for human review.

Usage:
    python scripts/fetch_news.py                 # fetch + write today's edition
    python scripts/fetch_news.py --dry-run       # fetch + print candidates, no writes, no LLM
    python scripts/fetch_news.py --max-stories 8 # cap the edition size (cost/length guard)
    python scripts/fetch_news.py --date 2026-07-13
"""

import re
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET

import requests
import openpyxl
import pandas as pd

import mif_common as mc

_RSS_HEADERS = {"User-Agent": "Mozilla/5.0 (MIF-news-bot; +https://github.com)"}


# ─── ARGS ─────────────────────────────────────────────────────────────────────
def parse_args(argv):
    opts = {"dry_run": False, "max_stories": 10, "max_candidates": 60,
            "date": mc.today_iso()}
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--dry-run":
            opts["dry_run"] = True; i += 1
        elif a == "--max-stories" and i + 1 < len(argv):
            opts["max_stories"] = int(argv[i + 1]); i += 2
        elif a == "--date" and i + 1 < len(argv):
            opts["date"] = argv[i + 1]; i += 2
        else:
            i += 1
    return opts


# ─── COMPANY LIST (same prospects file the portal uses) ───────────────────────
def load_company_names(limit=40):
    """Return the most strategic company names to build news queries from.
    Reads the same Master Database sheet the generator reads."""
    try:
        df = pd.read_excel(mc.PROSPECTS_FILE, sheet_name=mc.PROSPECTS_SHEET)
    except Exception as e:
        mc.log(f"Could not read prospects for company queries: {e}")
        return []
    names = []
    # Prefer high-priority / customer rows so queries stay on the names that matter.
    if "Priority Tier" in df.columns:
        df = df.sort_values("Priority Tier", na_position="last")
    for v in df[mc.COL_COMPANY_NAME].tolist():
        if mc.is_empty(v):
            continue
        s = str(v).strip()
        if s.lower() == "company name":
            continue
        names.append(s)
        if len(names) >= limit:
            break
    return names


# ─── RSS FETCH ────────────────────────────────────────────────────────────────
def google_news_rss(query):
    q = urllib.parse.quote(query)
    return f"https://news.google.com/rss/search?q={q}&hl=en-IN&gl=IN&ceid=IN:en"


def fetch_feed(url):
    """Fetch and parse an RSS/Atom feed with the stdlib (no feedparser dep).
    Returns a list of {title, link, source, published} dicts."""
    try:
        resp = requests.get(url, headers=_RSS_HEADERS, timeout=20)
        resp.raise_for_status()
        root = ET.fromstring(resp.content)
    except Exception as e:
        mc.log(f"feed error {url[:60]}: {e}")
        return []

    def text(el, tag):
        child = el.find(tag)
        return (child.text or "").strip() if child is not None and child.text else ""

    entries = []
    # RSS 2.0: channel/item ; Atom: entry
    items = root.findall(".//item")
    if items:
        for it in items:
            src_el = it.find("source")
            entries.append({
                "title": text(it, "title"),
                "link": text(it, "link"),
                "source": (src_el.text or "").strip() if src_el is not None and src_el.text else "",
                "published": text(it, "pubDate"),
                "summary": _strip_html(text(it, "description")),
            })
    else:
        ns = "{http://www.w3.org/2005/Atom}"
        for it in root.findall(f".//{ns}entry"):
            link_el = it.find(f"{ns}link")
            entries.append({
                "title": text(it, f"{ns}title"),
                "link": link_el.get("href", "") if link_el is not None else "",
                "source": "",
                "published": text(it, f"{ns}updated"),
                "summary": _strip_html(text(it, f"{ns}summary")),
            })
    return entries


_TAG_RE = re.compile(r"<[^>]+>")


def _strip_html(s):
    """Google News descriptions are HTML fragments; reduce to plain text."""
    if not s:
        return ""
    txt = _TAG_RE.sub(" ", s)
    txt = (txt.replace("&nbsp;", " ").replace("&amp;", "&")
              .replace("&lt;", "<").replace("&gt;", ">").replace("&#39;", "'")
              .replace("&quot;", '"'))
    return re.sub(r"\s+", " ", txt).strip()


def gather_candidates(opts):
    """Fetch + de-dup candidate stories. Returns (new_candidates, seen_set)."""
    seen, _ = mc.load_seen()
    queries = list(mc.NEWS_TOPIC_QUERIES)
    for name in load_company_names():
        queries.append(f'"{name}" steel OR expansion OR plant OR capex')

    urls = [google_news_rss(q) for q in queries] + list(mc.INDUSTRY_RSS)
    candidates, batch_keys = [], set()
    for url in urls:
        for e in fetch_feed(url):
            title = e["title"].strip()
            link = e["link"].strip()
            if not title or not link:
                continue
            if not mc.mark_seen(seen, title, link):      # already published
                continue
            k = mc._story_key(title, link)
            if k in batch_keys:                          # dup within this run
                continue
            batch_keys.add(k)
            candidates.append({
                "title": title, "link": link, "source": e.get("source", ""),
                "published": e.get("published", ""), "summary": e.get("summary", ""),
            })
        time.sleep(0.2)                                  # be polite to feeds
    candidates = candidates[: opts["max_candidates"]]
    mc.log(f"{len(candidates)} new candidate stories after de-dup")
    return candidates, seen


# ─── FX INDICATOR ─────────────────────────────────────────────────────────────
def fetch_usd_inr():
    try:
        r = requests.get(mc.FX_URL, timeout=15)
        r.raise_for_status()
        rate = float(r.json()["rates"]["INR"])
        return round(rate, 2)
    except Exception as e:
        mc.log(f"FX fetch failed: {e}")
        return None


# ─── SEGMENT / STATUS KEYWORD RULES (shared by LLM and rule-based modes) ──────
# Ordered: first match wins. Each entry -> (section, segment).
_SEGMENT_RULES = [
    (("hr coil", "hrc", "hot-rolled", "coil price", "steel price", "steel cos",
      "steelmaker", "steel prices", "scrap", "iron ore", "safeguard", "billet",
      "rebar", "import duty", "steel export", "steel import"), ("Input Cost", "Steel")),
    (("elevator", "lift", "escalator"), ("Segment", "Elevator")),
    (("solar", "module mounting", "renewable", "power"), ("Segment", "Power / Solar")),
    (("tata motors", "maruti", "mahindra", "automobile", "auto ", "vehicle",
      "ev ", "two-wheeler", "commercial vehicle", "oem"), ("Customer Health", "Automotive")),
    (("construction", "infrastructure", "real estate", "building", "cement"),
     ("Segment", "Construction")),
    (("tube", "pipe"), ("Segment", "Tubes & Pipes")),
    (("roll forming", "roll-forming", "cold roll"), ("Competitor", "Roll Forming")),
]
_HIGH_WORDS = ("hike", "surge", "shutdown", "closure", "crisis", "record",
               "safeguard", "layoff", "dumping", "ban", "high", "cut", "slump")
_PRICE_RE = re.compile(r"₹\s?\d[\d,]{3,}")


def _classify(text):
    t = text.lower()
    for kws, (section, segment) in _SEGMENT_RULES:
        if any(k in t for k in kws):
            return section, segment
    return "Update", "—"


def _status_for(text):
    return "High" if any(w in text.lower() for w in _HIGH_WORDS) else "Medium"


# ─── LLM SHAPING (free provider) ──────────────────────────────────────────────
_SYSTEM = (
    "You are the market-intelligence editor for Mother India Forming (MIF), an "
    "Indian roll-forming and steel-component manufacturer, briefing MIF's BD and "
    "procurement teams. Summarise ONLY from the headline text provided below — do "
    "NOT invent facts, numbers, or events not present in that text. Keep each "
    "story's real source URL. Be terse and factual. Output valid JSON only."
)


def build_prompt(candidates, date, max_stories):
    lines = [f"- TITLE: {c['title']}\n  SUMMARY: {c.get('summary','')}\n  "
             f"SOURCE: {c['source']}  URL: {c['link']}" for c in candidates]
    headlines = "\n".join(lines) if lines else "(no fresh headlines today)"
    return f"""Today is {date}. Below are fresh, de-duplicated news items for the Indian
steel / roll-forming / auto-component / construction market. Select the items that
genuinely matter to MIF (steel input costs, customer/OEM health, competitor moves,
demand in MIF's segments, new prospects) and shape them — using ONLY the text
given — into JSON with exactly this shape:
{{
  "briefing": {{"type": "Daily", "note": "<3-5 sentence procurement/BD briefing on steel input costs and what to watch>"}},
  "indicators": [
    {{"name": "HR Coil", "value": "<e.g. Rs 54,800-60,450>", "unit": "/ton",
      "movement": "Up|Down|Stable", "driver": "<short driver>", "source": "<publisher + url>"}}
  ],
  "stories": [
    {{"section": "Segment|Customer Health|Competitor|Input Cost|Prospect",
      "segment": "<e.g. Automotive/Construction/Elevator>",
      "status": "High|Medium|Low",
      "company": "<company or —>",
      "headline": "<one-line headline>",
      "what": "<2-3 sentence what happened, from the summary>",
      "why": "<why it matters to MIF, 1-2 sentences>",
      "source": "<publisher name + url from the item>",
      "importance": "High|normal",
      "marquee": "Y or blank (Y = show in the scrolling ticker; use sparingly)",
      "dashboard": "Y or blank (Y = surface on the dashboard)"}}
  ]
}}

Keep at most {max_stories} stories, ranked by importance to MIF. Only add an
indicator if a price/number appears in the text. News items:
{headlines}
"""


def shape_edition_llm(candidates, date, max_stories):
    """Free-LLM path (groq/gemini). Returns dict or None on any failure."""
    prompt = build_prompt(candidates, date, max_stories)
    try:
        data = mc.llm_json(prompt, system=_SYSTEM, max_tokens=6000)
    except mc.LLMUnavailable as e:
        mc.log(f"LLM unavailable ({e})")
        return None
    except Exception as e:
        mc.log(f"LLM shaping failed: {e}")
        return None
    return data if isinstance(data, dict) else None


# ─── RULE-BASED SHAPING (zero-key fallback) ───────────────────────────────────
def rule_based_edition(candidates, date, max_stories):
    """Build an edition straight from RSS with keyword rules — no LLM, no key.
    'Why It Matters' is intentionally left blank for the human to complete in the
    review PR. Every story keeps its real source URL."""
    # Rank: price/steel items and 'High' signal words first.
    def score(c):
        blob = f"{c['title']} {c.get('summary','')}"
        s = 0
        if _PRICE_RE.search(blob): s += 3
        if _status_for(blob) == "High": s += 2
        sect, _ = _classify(blob)
        if sect != "Update": s += 1
        return s
    ranked = sorted(candidates, key=score, reverse=True)[:max_stories]

    stories, indicators = [], []
    for c in ranked:
        blob = f"{c['title']} {c.get('summary','')}"
        section, segment = _classify(blob)
        src = (f"{c['source']} — " if c['source'] else "") + c['link']
        stories.append({
            "section": section, "segment": segment,
            "status": _status_for(blob), "company": "—",
            "headline": c['title'],
            "what": c.get('summary', '') or c['title'],
            "why": "",                                   # human fills in the PR
            "source": src,
            "importance": "High" if _status_for(blob) == "High" else "normal",
            "marquee": "", "dashboard": ""})
        # Capture a price mention as an HR-coil indicator when present.
        m = _PRICE_RE.search(blob)
        if m and any(k in blob.lower() for k in ("hr coil", "hrc", "hot-rolled", "steel")):
            indicators.append({
                "name": "HR Coil", "value": m.group(0).strip(), "unit": "/ton",
                "movement": "Up" if "hike" in blob.lower() or "surge" in blob.lower() else "Stable",
                "driver": c['title'][:80], "source": src})

    note = ("Auto-drafted from market headlines (no AI summary configured). "
            "Review the stories below and add the 'Why it matters to MIF' notes "
            "before publishing.")
    return {"briefing": {"type": "Daily", "note": note},
            "indicators": indicators[:1], "stories": stories}


def shape_edition(candidates, date, max_stories):
    """Choose the free path: LLM if a provider is configured, else rule-based."""
    if mc.llm_available():
        mc.log(f"Shaping edition with LLM provider: {mc.resolve_provider()}")
        data = shape_edition_llm(candidates, date, max_stories)
        if data:
            return data
        mc.log("LLM path failed — falling back to rule-based drafting.")
    else:
        mc.log("No LLM provider configured — using rule-based drafting.")
    return rule_based_edition(candidates, date, max_stories)


# ─── EXCEL WRITE ──────────────────────────────────────────────────────────────
def open_or_create_workbook():
    import os
    if os.path.exists(mc.NEWS_FILE):
        return openpyxl.load_workbook(mc.NEWS_FILE)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet, header in mc.NEWS_SHEETS.items():
        ws = wb.create_sheet(sheet)
        ws.append(header)
    return wb


def _ws_header(ws, expected):
    """Return the sheet's header row (create it if the sheet is empty)."""
    if ws.max_row == 0 or all(c.value is None for c in ws[1]):
        ws.append(expected)
        return list(expected)
    return [c.value for c in ws[1]]


def append_rows(ws, header, rows):
    hdr = _ws_header(ws, header)
    for row in rows:
        ws.append([row.get(col, "") for col in hdr])


def _yn(v):
    return "Y" if str(v).strip().lower() in ("y", "yes", "true", "1") else ""


def write_edition(edition, date, usd_inr):
    wb = open_or_create_workbook()

    # Briefing
    brief = edition.get("briefing") or {}
    if brief.get("note"):
        append_rows(wb["Briefing"], mc.NEWS_SHEETS["Briefing"], [{
            "Date": date, "Type": brief.get("type", "Daily"),
            "Input Cost Note": brief["note"]}])

    # Indicators
    ind_rows, hr_coil_val = [], None
    for ind in edition.get("indicators", []):
        ind_rows.append({
            "Date": date, "Indicator": ind.get("name", ""),
            "Value": ind.get("value", ""), "Unit": ind.get("unit", ""),
            "Movement": ind.get("movement", ""), "Driver": ind.get("driver", ""),
            "Source": ind.get("source", "")})
        if str(ind.get("name", "")).strip().lower().startswith("hr coil"):
            m = ''.join(ch for ch in str(ind.get("value", "")) if ch.isdigit())
            if m:
                hr_coil_val = float(m[:5]) if len(m) >= 5 else float(m)
    if ind_rows:
        append_rows(wb["Indicators"], mc.NEWS_SHEETS["Indicators"], ind_rows)

    # Trend (numeric point for the charts)
    trend_row = {"Date": date, "HR Coil": hr_coil_val if hr_coil_val else "",
                 "USD/INR": usd_inr if usd_inr else ""}
    if trend_row["HR Coil"] != "" or trend_row["USD/INR"] != "":
        append_rows(wb["Trend"], mc.NEWS_SHEETS["Trend"], [trend_row])

    # News stories
    story_rows = []
    for s in edition.get("stories", []):
        if not s.get("headline") or not s.get("source"):
            continue
        story_rows.append({
            "Date": date, "Section": s.get("section", "Update"),
            "Segment": s.get("segment", "—"), "Status": s.get("status", ""),
            "Company": s.get("company", "—"), "Headline": s["headline"],
            "What Happened": s.get("what", ""), "Why It Matters": s.get("why", ""),
            "Source": s["source"], "Importance": (s.get("importance") or "normal"),
            "Marquee": _yn(s.get("marquee")), "Dashboard": _yn(s.get("dashboard"))})
    if story_rows:
        append_rows(wb["News"], mc.NEWS_SHEETS["News"], story_rows)

    wb.save(mc.NEWS_FILE)
    return len(story_rows), len(ind_rows)


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    opts = parse_args(sys.argv[1:])
    mc.log(f"fetch_news — date={opts['date']} provider={mc.resolve_provider()} "
           f"dry_run={opts['dry_run']}")

    candidates, seen = gather_candidates(opts)
    usd_inr = fetch_usd_inr()
    mc.log(f"USD/INR = {usd_inr}")

    if opts["dry_run"]:
        print(f"\n── Candidate stories ({mc.resolve_provider()} mode) ──")
        for c in candidates:
            print(f"  • {c['title']}\n      {c['source']} — {c['link']}")
        print(f"\nUSD/INR indicator: {usd_inr}")
        print(f"\n{len(candidates)} candidates, cap {opts['max_stories']} stories/edition."
              " No LLM call, no Excel write (dry-run).")
        return

    if not candidates:
        mc.log("No new candidate stories today — nothing to write.")
        return

    edition = shape_edition(candidates, opts["date"], opts["max_stories"])
    n_stories, n_ind = write_edition(edition, opts["date"], usd_inr)
    mc.save_seen(seen)
    mc.log(f"Wrote {n_stories} stories, {n_ind} indicators to {mc.NEWS_FILE}")
    mc.log("Seen cache updated. Review the PR before it deploys.")


if __name__ == "__main__":
    main()
