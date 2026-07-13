#!/usr/bin/env python3
"""
discover_companies.py — find companies NOT yet in the Master Database and add
them as new rows (all 56 columns present), flagged for human verification.

Discovery sources (all free):
  1. News-mine  — org-like names pulled from the day's news headlines/summaries.
  2. Exhibitors — "Key Customers Exhibiting" / "Competitor Presence" names in the
     exhibitions workbook (parsed by generate_portal_v7_News.read_exhibitions()).
  3. Wikidata   — best-effort SPARQL for Indian companies by industry per segment.

A candidate is only ADDED if it (a) is not already in the DB (normalised match),
(b) is not already in data/discovered_seen.json, and (c) can be confirmed in a
free structured source (Wikidata/GLEIF) — this keeps out random capitalised
phrases and unfindable names.

New rows are written with identity fields filled, everything else blank, and
Estimated vs Verified Flag = "Needs Verification" — nothing is auto-"verified".
A human promotes rows via the review PR.

Usage:
    python scripts/discover_companies.py --dry-run --limit 10
    python scripts/discover_companies.py --limit 10
"""

import re
import sys
import importlib.util

import openpyxl
import requests

import mif_common as mc
import mif_context as ctx
import enrich_data as ed
import fetch_news as fn

_HTTP = {"User-Agent": "MIF-portal-discovery/1.0"}


def parse_args(argv):
    opts = {"dry_run": False, "limit": 10, "max_candidates": 120}
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--dry-run":
            opts["dry_run"] = True; i += 1
        elif a == "--limit" and i + 1 < len(argv):
            opts["limit"] = int(argv[i + 1]); i += 2
        else:
            i += 1
    return opts


# ─── CANDIDATE GATHERING ──────────────────────────────────────────────────────
# Org-like: a run of Capitalised tokens containing/ending in a company marker.
_ORG_MARKERS = (r"Ltd|Limited|Pvt|Private|LLP|Industries|Engineering|Motors|"
                r"Steel|Steels|Corporation|Corp|Enterprises|Manufacturing|"
                r"Fabricators|Fabrication|Tubes|Techno|Auto|Metals?|Alloys|"
                r"Structurals?|Systems|Technologies|Infra")
_ORG_RE = re.compile(
    r"\b([A-Z][A-Za-z&.\-]+(?:\s+[A-Z][A-Za-z&.\-]+){0,4}\s+(?:" + _ORG_MARKERS + r"))\b")


def _clean_candidate(name):
    return re.sub(r"\s+", " ", name.strip(" .,-&")).strip()


def from_news():
    """Extract org-like names from today's news candidates. (source label, name)."""
    out = []
    cands, _ = fn.gather_candidates({"max_candidates": 60})
    for c in cands:
        blob = f"{c['title']} {c.get('summary','')}"
        for m in _ORG_RE.findall(blob):
            nm = _clean_candidate(m)
            if nm:
                out.append(("news", nm, blob))
    return out


def from_exhibitors():
    """Split exhibitor / competitor name lists from the exhibitions workbook."""
    spec = importlib.util.spec_from_file_location(
        "gen", str(mc.REPO_ROOT) + "/generate_portal_v7_News.py")
    gen = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(gen)
    out = []
    for e in gen.read_exhibitions():
        seg_hint = e.get("segment", "")
        for field in ("keyCustomers", "competitors"):
            for nm in re.split(r"[,/;]", str(e.get(field, "") or "")):
                nm = _clean_candidate(nm)
                if nm and len(nm.split()) <= 6:
                    out.append(("exhibitor", nm, f"{seg_hint} {nm}"))
    return out


_SPARQL = """SELECT ?cLabel WHERE {{
  ?c wdt:P31/wdt:P279* wd:Q4830453 ; wdt:P17 wd:Q668 ; wdt:P452 wd:{qid} .
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en". }}
}} LIMIT 15"""

# A small, best-effort segment -> industry Qid map (Wikidata P452 values). Only
# reliably-mapped industries are listed; other segments rely on news/exhibitor
# mining. Extend as you confirm more industry Qids.
_SEGMENT_INDUSTRY_QID = {
    "Auto Components": "Q190117",            # automotive industry
    "Auto - Passenger (LMV)": "Q190117",
    "Auto - Commercial (LCV_MCV)": "Q190117",
    "Bus Body": "Q190117",
}


def from_wikidata():
    """Best-effort: Indian companies by industry per mapped segment."""
    out = []
    for seg, qid in _SEGMENT_INDUSTRY_QID.items():
        if not re.fullmatch(r"Q\d+", qid):
            continue
        try:
            r = requests.get("https://query.wikidata.org/sparql",
                             params={"query": _SPARQL.format(qid=qid), "format": "json"},
                             headers=_HTTP, timeout=25)
            r.raise_for_status()
            for b in r.json().get("results", {}).get("bindings", []):
                nm = _clean_candidate(b.get("cLabel", {}).get("value", ""))
                if nm and not re.fullmatch(r"Q\d+", nm):
                    out.append(("wikidata", nm, seg))
        except Exception as e:
            mc.log(f"  wikidata sparql ({seg}) skipped: {e}")
    return out


def gather_candidates(opts):
    raw = from_news() + from_exhibitors() + from_wikidata()
    # De-dup within this run by normalised name, keep first (source, name, ctx).
    seen_norm, uniq = set(), []
    for source, name, context in raw:
        if not mc.looks_like_company(name):
            continue
        nk = mc.normalize_name(name)
        if not nk or nk in seen_norm:
            continue
        seen_norm.add(nk)
        uniq.append((source, name, context))
    return uniq[: opts["max_candidates"]]


# ─── IDENTITY CONFIRMATION (strict — avoids fuzzy false positives) ────────────
def _name_match(candidate, official):
    """True if the candidate name is fully contained in the official name
    (normalised tokens), so GLEIF's fuzzy legalName filter can't slip an
    unrelated record through."""
    ct = set(mc.normalize_name(candidate).split())
    ot = set(mc.normalize_name(official).split())
    return bool(ct) and ct <= ot


def confirm_identity(name):
    """Confirm a candidate is a real company via free structured sources and
    return {field: {value, source}}. Requires EITHER a Wikidata business entity
    OR a GLEIF record whose legal name actually matches — otherwise {}.
    """
    wd = ed._wikidata_facts(name)
    gl = ed._gleif_facts(name)
    gl_ok = bool(gl) and _name_match(name, gl.get("_name", ""))
    if not wd and not gl_ok:
        return {}

    out = {}
    for field, key in (("Company Website", "website"), ("Year Est.", "year"),
                       ("Parent / Group", "parent"), ("Products", "industry"),
                       ("Legal Form", "legalform")):
        if wd.get(key):
            out[field] = {"value": wd[key], "source": wd["_source"]}
    # HQ: prefer a verified GLEIF registered address, else Wikidata's city.
    if gl_ok and gl.get("hq"):
        out["HQ Address"] = {"value": gl["hq"], "source": gl["_source"]}
    elif wd.get("hq"):
        out["HQ Address"] = {"value": wd["hq"], "source": wd["_source"]}
    return out


# ─── ROW BUILDING ─────────────────────────────────────────────────────────────
def _next_sno(ws, sno_col):
    mx = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, sno_col).value
        try:
            mx = max(mx, int(float(v)))
        except (TypeError, ValueError):
            pass
    return mx + 1


def build_row(headers, name, segment, identity, sources, source_label, sno):
    """Return a full 56-column ordered list for one new company."""
    d = {h: "" for h in headers}
    d["Company Name"] = name
    d["Segment"] = segment or "Uncategorized"
    if "S.No" in d:
        d["S.No"] = sno
    for field, payload in identity.items():
        if field in d:
            d[field] = payload["value"]
    d["Relationship Status"] = "Not Contacted"
    d["Priority Tier"] = "Monitor"
    d["Competitor Tier"] = "Pure Prospect"
    d["Data Confidence"] = "Low"
    d[mc.COL_EST_FLAG] = "Needs Verification"
    d["Verification Method"] = "Auto-discovered"
    d["Last Verified Date"] = mc.today_iso()
    d[mc.COL_SOURCE_LINKS] = "\n".join(sources)
    d["BD Notes"] = (f"Auto-discovered from {source_label} on {mc.today_iso()} — "
                     "identity from free sources; verify before outreach.")
    return [d.get(h, "") for h in headers]


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    opts = parse_args(sys.argv[1:])
    mc.log(f"discover_companies — limit={opts['limit']} dry_run={opts['dry_run']}")

    db_index = mc.db_name_index()
    seen, _ = mc.load_discovered()
    candidates = gather_candidates(opts)
    mc.log(f"{len(candidates)} unique candidate names gathered "
           f"(news + exhibitors + wikidata).")

    # Keep only genuinely-new names.
    fresh = [(s, n, cx) for s, n, cx in candidates
             if mc.normalize_name(n) not in db_index
             and mc.normalize_name(n) not in seen]
    mc.log(f"{len(fresh)} are new (not in DB, not previously discovered).")

    wb = openpyxl.load_workbook(mc.PROSPECTS_FILE)
    ws = wb[mc.PROSPECTS_SHEET]
    headers = [c.value for c in ws[1]]
    hmap = ed.header_map(ws)
    sno_col = hmap.get("S.No")

    added, checked = 0, 0
    preview = []
    for source, name, context in fresh:
        if added >= opts["limit"]:
            break
        checked += 1
        # Relevance gate: must classify into one of MIF's end-use segments. This
        # drops steel producers / off-target names mined from price news.
        segment = ctx.normalize_segment(ctx.classify_segment(context))
        if not segment and source == "wikidata":
            segment = ctx.normalize_segment(context)              # the queried segment
        if not segment:
            continue
        # Existence gate: confirmed real company in a free structured source.
        identity = confirm_identity(name)
        if not identity:
            continue
        sources = sorted({p["source"] for p in identity.values()})
        preview.append((source, name, segment or "Uncategorized",
                        list(identity.keys()), sources))
        if not opts["dry_run"]:
            sno = _next_sno(ws, sno_col) if sno_col else ""
            ws.append(build_row(headers, name, segment, identity, sources, source, sno))
            wa = ed.ensure_audit_sheet(wb)
            ed.audit(wa, name, "New company",
                     f"Added ({segment or 'Uncategorized'})", "; ".join(sources))
        seen.add(mc.normalize_name(name))
        added += 1

    if opts["dry_run"]:
        print(f"\n── {len(preview)} companies WOULD be added "
              f"(checked {checked} candidates) ──")
        for src, name, seg, fields, srcs in preview:
            print(f"  • {name}  [{seg}]  via {src}")
            print(f"      fields: {', '.join(fields)}")
            print(f"      source: {srcs[0] if srcs else '-'}")
        print("\nDry-run: no rows written.")
        return

    if added:
        wb.save(mc.PROSPECTS_FILE)
        mc.save_discovered(seen)
        mc.log(f"Added {added} new companies to {mc.PROSPECTS_SHEET} "
               "(flagged 'Needs Verification'). Review the PR before deploy.")
    else:
        mc.log("No confirmable new companies this run — nothing written.")


if __name__ == "__main__":
    main()
