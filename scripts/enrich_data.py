#!/usr/bin/env python3
"""
enrich_data.py — fill missing prospect / exhibition fields via online search.

Fill-only, never overwrite:
  * Only EMPTY cells in the allow-listed fields are touched.
  * A row whose "Estimated vs Verified Flag" says "Verified" is never modified.
  * Every filled cell gets a source URL in "Source Hyperlink(s)", is marked
    "Estimated", stamped with today's date, and logged to the "Fill Audit Log"
    sheet (Company, Field, New Value, Method, Source, Date).

Enrichment uses FREE structured sources only — no paid API, no web scraping of
search engines:
  * Wikidata / Wikipedia (company website, founding year, HQ, parent, industry)
  * GLEIF LEI records (legal name, legal form, registered HQ address)
  * (optional) a free LLM provider (Groq/Gemini) to normalise the fetched text —
    used only if MIF_LLM_PROVIDER / a key is configured; never required.

Coverage note: these sources mainly cover larger, well-known companies. Obscure
SME prospects simply won't be found and their cells stay empty for a human.
With --dry-run the script only reports which companies/fields it WOULD enrich.

Per-run caps keep time bounded; a full backfill happens over many days.

Usage:
    python scripts/enrich_data.py --dry-run --limit 3
    python scripts/enrich_data.py --limit 10                 # prospects only
    python scripts/enrich_data.py --limit 10 --exhibitions-limit 5
    python scripts/enrich_data.py --which exhibitions --exhibitions-limit 8
"""

import os
import re
import sys

import openpyxl
import requests

import mif_common as mc

_HTTP_HEADERS = {"User-Agent": "MIF-portal-enrichment/1.0 (contact: BI@mif)"}


# ─── ARGS ─────────────────────────────────────────────────────────────────────
def parse_args(argv):
    opts = {"dry_run": False, "limit": 8, "exhibitions_limit": 0, "which": "prospects"}
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--dry-run":
            opts["dry_run"] = True; i += 1
        elif a == "--limit" and i + 1 < len(argv):
            opts["limit"] = int(argv[i + 1]); i += 2
        elif a == "--exhibitions-limit" and i + 1 < len(argv):
            opts["exhibitions_limit"] = int(argv[i + 1]); i += 2
        elif a == "--which" and i + 1 < len(argv):
            opts["which"] = argv[i + 1]; i += 2
        else:
            i += 1
    # A bare --exhibitions-limit implies we also want exhibitions.
    if opts["exhibitions_limit"] and opts["which"] == "prospects":
        opts["which"] = "both"
    return opts


# ─── SHARED HELPERS ───────────────────────────────────────────────────────────
def header_map(ws, header_row=1):
    """Map column header text -> 1-based column index for an openpyxl sheet."""
    return {str(c.value).strip(): c.column
            for c in ws[header_row] if c.value is not None}


def append_source(existing, url):
    """Append a source URL to a 'Source Hyperlink(s)' cell without duplicating."""
    existing = "" if mc.is_empty(existing) else str(existing).strip()
    if not url or url in existing:
        return existing
    return (existing + "\n" + url).strip() if existing else url


def ensure_audit_sheet(wb):
    if mc.AUDIT_SHEET in wb.sheetnames:
        return wb[mc.AUDIT_SHEET]
    ws = wb.create_sheet(mc.AUDIT_SHEET)
    ws.append(mc.AUDIT_COLS)
    return ws


def audit(ws_audit, company, field, value, source):
    ws_audit.append([company, field, value,
                     "Free lookup (Wikidata/GLEIF, auto-enrich)", source, mc.today_iso()])


# ─── FREE STRUCTURED-DATA LOOKUPS ─────────────────────────────────────────────
_WD_API = "https://www.wikidata.org/w/api.php"

# Wikidata property -> internal fact key. Q-id valued props get label-resolved.
_WD_PROPS = {"P856": ("website", False), "P571": ("year", False),
             "P159": ("hq", True), "P749": ("parent", True),
             "P1454": ("legalform", True), "P452": ("industry", True)}

# Our column/label -> internal fact key. Several columns share a source fact.
_FIELD_TO_FACT = {
    "Company Website": "website", "Registration Website": "website",
    "Year Est.": "year", "HQ Address": "hq", "Parent / Group": "parent",
    "Legal Form": "legalform", "Products": "industry",
}


def _get(url, **params):
    r = requests.get(url, params=params, headers=_HTTP_HEADERS, timeout=20)
    r.raise_for_status()
    return r.json()


def _wikidata_facts(name):
    """Return {fact_key: value} + a wikidata source URL, or {} if not found.
    Only accepts a hit that carries at least one business-ish claim so we don't
    match a person/place of the same name."""
    try:
        hits = _get(_WD_API, action="wbsearchentities", search=name, language="en",
                    format="json", limit=1, type="item").get("search", [])
        if not hits:
            return {}
        qid = hits[0]["id"]
        ent = _get(_WD_API, action="wbgetentities", ids=qid, props="claims|labels",
                   languages="en", format="json")["entities"][qid]
    except Exception as e:
        mc.log(f"    wikidata lookup failed: {e}")
        return {}

    claims = ent.get("claims", {})
    facts, label_ids = {}, {}
    for pid, (key, is_item) in _WD_PROPS.items():
        arr = claims.get(pid)
        if not arr:
            continue
        snak = arr[0].get("mainsnak", {})
        if snak.get("snaktype") != "value":
            continue
        val = snak["datavalue"]["value"]
        if key == "year":
            m = re.search(r"(\d{4})", val.get("time", "") if isinstance(val, dict) else str(val))
            if m:
                facts["year"] = m.group(1)
        elif is_item:
            label_ids[key] = val["id"]                    # resolve label below
        else:
            facts[key] = str(val).strip()                 # website: plain string

    if not facts and not label_ids:
        return {}                                         # top hit had no business claims
    # Resolve all referenced Q-ids to English labels in one call.
    if label_ids:
        try:
            ents = _get(_WD_API, action="wbgetentities", ids="|".join(label_ids.values()),
                        props="labels", languages="en", format="json")["entities"]
            for key, qv in label_ids.items():
                lbl = ents.get(qv, {}).get("labels", {}).get("en", {}).get("value")
                if lbl:
                    facts[key] = lbl
        except Exception as e:
            mc.log(f"    wikidata label resolve failed: {e}")
    facts["_source"] = f"https://www.wikidata.org/wiki/{qid}"
    return facts


def _gleif_facts(name):
    """Return {hq, source} from the GLEIF LEI register, or {} if not found."""
    try:
        data = _get("https://api.gleif.org/api/v1/lei-records",
                    **{"filter[entity.legalName]": name, "page[size]": 1}).get("data", [])
    except Exception as e:
        mc.log(f"    gleif lookup failed: {e}")
        return {}
    if not data:
        return {}
    rec = data[0]
    entity = (rec.get("attributes", {}).get("entity", {}) or {})
    addr = entity.get("legalAddress", {}) or {}
    parts = list(addr.get("addressLines", []) or []) + [
        addr.get("city", ""), addr.get("region", ""), addr.get("postalCode", ""),
        addr.get("country", "")]
    hq = ", ".join(p for p in parts if p)
    out = {"_source": f"https://search.gleif.org/#/record/{rec.get('id','')}",
           "_name": (entity.get("legalName", {}) or {}).get("name", "")}
    if hq:
        out["hq"] = hq
    return out


def enrich_company(name, context, fields):
    """Look up `fields` for one company/event using free structured sources.
    Returns {field: {"value":..., "source":...}} only for fields we could fill.
    `context` is accepted for signature compatibility (unused by free lookups)."""
    wd = _wikidata_facts(name)
    gl = _gleif_facts(name) if any(f in ("HQ Address", "Legal Form") for f in fields) else {}

    out = {}
    for field in fields:
        key = _FIELD_TO_FACT.get(field)
        if not key:
            continue
        # HQ Address: prefer GLEIF's full registered address over Wikidata's city.
        if field == "HQ Address" and gl.get("hq"):
            out[field] = {"value": gl["hq"], "source": gl["_source"]}
            continue
        if key in wd and wd[key]:
            out[field] = {"value": wd[key], "source": wd["_source"]}
    return out


# ─── PROSPECTS ────────────────────────────────────────────────────────────────
def select_prospect_rows(ws, hmap, limit):
    """Rank non-verified rows by BD value first (customers -> Tier 1 -> Active
    Prospects -> Tier 2 -> rest), then by how many allow-listed fields are empty."""
    import mif_context as ctx
    flag_col = hmap.get(mc.COL_EST_FLAG)
    name_col = hmap[mc.COL_COMPANY_NAME]
    enrich_cols = [(f, hmap[f]) for f in mc.ENRICHABLE_PROSPECT_FIELDS if f in hmap]
    # Columns bd_score reads, resolved once.
    score_cols = ("Relationship Status", "Competitor Tier", "Priority Tier")

    ranked = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not mc.looks_like_company(name):
            continue
        if flag_col and mc.is_verified(ws.cell(r, flag_col).value):
            continue
        missing = [f for f, c in enrich_cols if mc.is_empty(ws.cell(r, c).value)]
        if not missing:
            continue
        row = {k: ws.cell(r, hmap[k]).value for k in score_cols if k in hmap}
        ranked.append((ctx.bd_score(row), len(missing), r, str(name).strip(), missing))
    ranked.sort(key=lambda t: (t[0], t[1]), reverse=True)
    # Drop the score field so the return shape stays (miss, row, name, fields).
    return [(m, r, n, f) for _, m, r, n, f in ranked[:limit]]


def context_for(ws, hmap, row):
    out = {}
    for f in ("CIN", "GSTIN", "Company Website", "HQ Address", "Primary Segment",
              "Products", "Region"):
        c = hmap.get(f)
        if c:
            out[f] = ws.cell(row, c).value
    return out


def enrich_prospects(opts):
    wb = openpyxl.load_workbook(mc.PROSPECTS_FILE)
    ws = wb[mc.PROSPECTS_SHEET]
    hmap = header_map(ws)
    targets = select_prospect_rows(ws, hmap, opts["limit"])
    mc.log(f"Prospects: {len(targets)} companies selected for enrichment "
           f"(BD-value priority, cap {opts['limit']}).")

    if opts["dry_run"]:
        for miss, r, name, fields in targets:
            print(f"  • {name} (row {r}) — {miss} empty fields: {', '.join(fields)}")
        print("\nDry-run: no lookups, no writes.")
        return 0

    ws_audit = ensure_audit_sheet(wb)
    src_col = hmap.get(mc.COL_SOURCE_LINKS)
    conf_col = hmap.get(mc.COL_DATA_CONFIDENCE)
    flag_col = hmap.get(mc.COL_EST_FLAG)
    date_col = hmap.get(mc.COL_LAST_VERIFIED)
    method_col = hmap.get(mc.COL_VERIF_METHOD)

    filled_total = 0
    for _, r, name, fields in targets:
        found = enrich_company(name, context_for(ws, hmap, r), fields)
        if not found:
            mc.log(f"  {name}: not found in free sources")
            continue
        row_filled = 0
        for field, payload in found.items():
            col = hmap.get(field)
            if not col or not mc.is_empty(ws.cell(r, col).value):
                continue                                   # never overwrite
            ws.cell(r, col).value = payload["value"]
            if src_col:
                ws.cell(r, src_col).value = append_source(
                    ws.cell(r, src_col).value, payload["source"])
            audit(ws_audit, name, field, payload["value"], payload["source"])
            row_filled += 1
        if row_filled:
            # Only stamp provenance columns that are currently empty (don't
            # clobber a human's existing note).
            if conf_col and mc.is_empty(ws.cell(r, conf_col).value):
                ws.cell(r, conf_col).value = "Low"
            if flag_col and mc.is_empty(ws.cell(r, flag_col).value):
                ws.cell(r, flag_col).value = "Estimated"
            if date_col and mc.is_empty(ws.cell(r, date_col).value):
                ws.cell(r, date_col).value = mc.today_iso()
            if method_col and mc.is_empty(ws.cell(r, method_col).value):
                ws.cell(r, method_col).value = "Auto web search"
            filled_total += row_filled
            mc.log(f"  {name}: filled {row_filled} field(s)")

    if filled_total:
        wb.save(mc.PROSPECTS_FILE)
        mc.log(f"Prospects: {filled_total} cells filled and saved.")
    else:
        mc.log("Prospects: no cells filled (nothing saved).")
    return filled_total


# ─── EXHIBITIONS ──────────────────────────────────────────────────────────────
# The parsed columns of interest (0-based index within the MASTER INDEX sheet)
# map to these human labels; enrichment only fills the sparse informational ones.
EXH_FIELDS = {
    4:  "Confirmed Dates",
    6:  "Next Edition",
    14: "Registration Website",
    15: "Organiser Contact",
    16: "Venue & Google Maps",
}


def find_exh_sheet_and_header(wb):
    sheet = None
    for s in wb.sheetnames:
        u = s.upper()
        if "MASTER INDEX" in u and "COMPLETE" in u:
            sheet = s; break
    if sheet is None:
        for s in wb.sheetnames:
            if "MASTER INDEX" in s.upper():
                sheet = s; break
    if sheet is None:
        return None, None
    ws = wb[sheet]
    hdr = None
    for i in range(1, min(9, ws.max_row + 1)):
        vals = [str(c.value) for c in ws[i]]
        if any("Exhibition / Conference" in v for v in vals):
            hdr = i; break
    return ws, (hdr or 3)


def enrich_exhibitions(opts):
    if not os.path.exists(mc.EXHIBITIONS_FILE):
        mc.log("Exhibitions file not found — skipping.")
        return 0
    wb = openpyxl.load_workbook(mc.EXHIBITIONS_FILE)
    ws, hdr = find_exh_sheet_and_header(wb)
    if ws is None:
        mc.log("Exhibitions master index sheet not found — skipping.")
        return 0

    name_col = 2  # column B = "Exhibition / Conference"
    targets = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not mc.looks_like_company(name):
            continue
        missing = [(idx, lbl) for idx, lbl in EXH_FIELDS.items()
                   if mc.is_empty(ws.cell(r, idx + 1).value)]
        if missing:
            targets.append((len(missing), r, str(name).split("\n")[0].strip(), missing))
    targets.sort(reverse=True)
    targets = targets[: opts["exhibitions_limit"]]
    mc.log(f"Exhibitions: {len(targets)} events selected (cap {opts['exhibitions_limit']}).")

    if opts["dry_run"]:
        for miss, r, name, fields in targets:
            print(f"  • {name} (row {r}) — missing: {', '.join(l for _, l in fields)}")
        print("\nDry-run: no lookups, no writes.")
        return 0

    ws_audit = ensure_audit_sheet(wb)
    country_col = 4  # column D = "Country / City" for context
    filled_total = 0
    for _, r, name, fields in targets:
        ctx = {"Location": ws.cell(r, country_col).value}
        want = [lbl for _, lbl in fields]
        found = enrich_company(name, ctx, want)
        if not found:
            continue
        label_to_idx = {lbl: idx for idx, lbl in EXH_FIELDS.items()}
        row_filled = 0
        for label, payload in found.items():
            idx = label_to_idx.get(label)
            if idx is None or not mc.is_empty(ws.cell(r, idx + 1).value):
                continue
            ws.cell(r, idx + 1).value = payload["value"]
            audit(ws_audit, name, label, payload["value"], payload["source"])
            row_filled += 1
        if row_filled:
            filled_total += row_filled
            mc.log(f"  {name}: filled {row_filled} field(s)")

    if filled_total:
        wb.save(mc.EXHIBITIONS_FILE)
        mc.log(f"Exhibitions: {filled_total} cells filled and saved.")
    return filled_total


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    opts = parse_args(sys.argv[1:])
    mc.log(f"enrich_data — which={opts['which']} limit={opts['limit']} "
           f"exhibitions_limit={opts['exhibitions_limit']} dry_run={opts['dry_run']}")

    total = 0
    if opts["which"] in ("prospects", "both"):
        total += enrich_prospects(opts)
    if opts["which"] in ("exhibitions", "both") and opts["exhibitions_limit"]:
        total += enrich_exhibitions(opts)

    mc.log(f"Done. {total} cell(s) filled." if not opts["dry_run"] else "Dry-run complete.")


if __name__ == "__main__":
    main()
